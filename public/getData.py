from openpyxl import load_workbook

def getData(filename, testPoint):
    wb = load_workbook('E:\\Projects\\bhAPItest\\data\\'+ filename)
    ws = wb['Sheet1']
    start = 0 #起始位置
    interval = 0#间隔位置
    for cell in ws['D']:
        start = start+1
        if (cell.value == testPoint):#直到找到testPoint确定起始位置
            data = ws['D'][start:] #将起始位置后剩余的数据用list保存起来
            for cell in data:
                 interval = interval + 1 #直到找到下一个不为None的数据时，间隔位置加1
                 if cell.value != None: #
                     interval = interval -1 #去掉不为None的
                     break
            break
    start =  start -1
    end = start + interval
    cases = [] #用来装每个用例的list
    for cellId in range(start, end+1):
        case = [] #每个用例的list
        case.append(ws['E'][cellId].value)
        case.append(ws['F'][cellId].value)
        case.append(ws['G'][cellId].value)
        cases.append(case)
    return tuple(cases)


if __name__ == "__main__":
    cases = getData('dealer_dealername_success')
    print(cases)





